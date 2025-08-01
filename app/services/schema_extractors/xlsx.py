import io
import json
from typing import List, Optional, Tuple
from datetime import datetime, date, time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from fastapi import UploadFile
from . import BaseSchemaExtractor, DataSourceSchema, TableSchema, ColumnSchema, DataType


class XLSXSchemaExtractor(BaseSchemaExtractor):
    """Excel-specific schema extractor using unified architecture"""
    
    def __init__(self, sample_size: Optional[int] = 1000):
        self.sample_size = sample_size
    
    async def extract_schema(self, upload_file: UploadFile, **kwargs) -> DataSourceSchema:
        """
        Extract unified schema from Excel file.
        
        Args:
            upload_file: FastAPI UploadFile containing Excel data
            **kwargs: Additional options (analyze_all_sheets, sheet_name, sample_size)
            
        Returns:
            DataSourceSchema: Unified schema representation
        """
        try:
            # Read file content
            file_content = await upload_file.read()
            await upload_file.seek(0)
            
            if not file_content:
                raise ValueError("Excel file is empty")
            
            # Parse kwargs
            analyze_all_sheets = kwargs.get('analyze_all_sheets', True)
            specific_sheet = kwargs.get('sheet_name')
            sample_size = kwargs.get('sample_size', self.sample_size)
            
            # Load workbook
            workbook = load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
            
            # Determine sheets to analyze
            sheets_to_analyze = self._get_sheets_to_analyze(
                workbook, analyze_all_sheets, specific_sheet
            )
            
            # Extract schema from each sheet
            tables = []
            for sheet_name in sheets_to_analyze:
                sheet = workbook[sheet_name]
                table_schema = self._analyze_sheet(sheet, sheet_name, sample_size)
                if table_schema:  # Only add non-empty sheets
                    tables.append(table_schema)
            
            workbook.close()
            
            # Calculate metadata
            total_rows = sum(table.row_count or 0 for table in tables)
            file_size_mb = len(file_content) / (1024 * 1024)
            
            # Determine business context
            business_context = self._infer_business_context(tables)
            
            return DataSourceSchema(
                source_name=upload_file.filename or "excel_file.xlsx",
                source_type="xlsx",
                tables=tables,
                metadata={
                    "file_size_mb": round(file_size_mb, 2),
                    "total_sheets": len(workbook.sheetnames),
                    "analyzed_sheets": len(tables),
                    "business_context": business_context,
                    "has_multiple_sheets": len(tables) > 1,
                    "extraction_sample_size": sample_size
                }
            )
            
        except Exception as e:
            raise Exception(f"Error extracting XLSX schema: {e}")
    
    def _get_sheets_to_analyze(self, workbook, analyze_all_sheets: bool, specific_sheet: Optional[str]) -> List[str]:
        """Determine which sheets to analyze"""
        if specific_sheet:
            if specific_sheet in workbook.sheetnames:
                return [specific_sheet]
            else:
                raise ValueError(f"Sheet '{specific_sheet}' not found. Available: {workbook.sheetnames}")
        elif analyze_all_sheets:
            return workbook.sheetnames
        else:
            return [workbook.sheetnames[0]] if workbook.sheetnames else []
    
    def _analyze_sheet(self, sheet, sheet_name: str, sample_size: Optional[int]) -> Optional[TableSchema]:
        """Analyze individual Excel sheet"""
        try:
            # Get actual data range
            max_row, max_col = self._get_actual_data_range(sheet)
            
            if max_row == 0 or max_col == 0:
                return None  # Empty sheet
            
            # Limit analysis if sample_size specified
            rows_to_analyze = min(sample_size, max_row) if sample_size else max_row
            
            # Extract headers (assume first row)
            headers = self._extract_headers(sheet, max_col)
            
            # Analyze each column
            columns = []
            for col_idx, header in enumerate(headers, 1):
                column_schema = self._analyze_column(sheet, col_idx, header, rows_to_analyze)
                columns.append(column_schema)
            
            # Determine table type and characteristics
            table_type = self._determine_table_type(sheet_name, columns)
            
            return TableSchema(
                name=sheet_name,
                columns=columns,
                row_count=max_row - 1,  # Subtract header row
                table_type=table_type,
                description=self._generate_table_description(sheet_name, columns, max_row)
            )
            
        except Exception as e:
            # Log warning but don't fail entire extraction
            print(f"Warning: Failed to analyze sheet '{sheet_name}': {e}")
            return None
    
    def _get_actual_data_range(self, sheet) -> Tuple[int, int]:
        """Get actual data range excluding empty rows/columns"""
        max_row = 0
        max_col = 0
        
        for row in sheet.iter_rows():
            row_num = row[0].row
            has_data = any(cell.value is not None for cell in row)
            if has_data:
                max_row = max(max_row, row_num)
                # Find rightmost column with data
                for col_idx, cell in enumerate(row, 1):
                    if cell.value is not None:
                        max_col = max(max_col, col_idx)
        
        return max_row, max_col
    
    def _extract_headers(self, sheet, max_col: int) -> List[str]:
        """Extract column headers from first row"""
        headers = []
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=1, column=col)
            header = str(cell.value).strip() if cell.value is not None else f"Column_{get_column_letter(col)}"
            headers.append(header)
        return headers
    
    def _analyze_column(self, sheet, col_idx: int, header: str, max_row: int) -> ColumnSchema:
        """Analyze individual Excel column"""
        # Collect cell data (skip header row)
        cells = []
        values = []
        
        for row in range(2, max_row + 1):  # Start from row 2 (skip header)
            cell = sheet.cell(row=row, column=col_idx)
            cells.append(cell)
            values.append(cell.value)
        
        # Basic statistics
        total_count = len(values)
        null_count = sum(1 for v in values if v is None)
        non_null_values = [v for v in values if v is not None]
        non_null_count = len(non_null_values)
        
        # Type analysis
        data_type, original_type = self._analyze_excel_column_type(cells, non_null_values)
        
        # Semantic type inference
        semantic_type = self._infer_semantic_type(header, data_type, [str(v) for v in non_null_values[:5]])
        
        # Sample values
        sample_values = [str(v) for v in non_null_values[:3] if v is not None]
        
        # Create base column schema
        column = ColumnSchema(
            name=header,
            data_type=semantic_type,
            original_type=original_type,
            is_nullable=null_count > 0,
            sample_values=sample_values,
            value_count=total_count,
            null_count=null_count,
            unique_count=len(set(str(v) for v in non_null_values)) if non_null_values else 0,
            description=self._generate_column_description(header, semantic_type, non_null_values)
        )
        
        # Add type-specific statistics
        self._add_type_specific_stats(column, non_null_values, semantic_type)
        
        # Detect constraints and relationships
        self._detect_column_constraints(column, non_null_values)
        
        return column
    
    def _analyze_excel_column_type(self, cells: List, values: List) -> Tuple[DataType, str]:
        """Analyze Excel column types"""
        if not values:
            return DataType.UNKNOWN, "empty"
        
        # Count types
        type_counts = {
            DataType.INTEGER: 0,
            DataType.DECIMAL: 0,
            DataType.TEXT: 0,
            DataType.BOOLEAN: 0,
            DataType.DATE: 0,
            DataType.DATETIME: 0,
            DataType.TIME: 0
        }
        
        for cell in cells:
            if cell.value is None:
                continue
                
            value = cell.value
            
            # Check Excel cell types and Python types
            if isinstance(value, bool):
                type_counts[DataType.BOOLEAN] += 1
            elif isinstance(value, int):
                type_counts[DataType.INTEGER] += 1
            elif isinstance(value, float):
                # Check if it's actually an integer stored as float
                if value.is_integer():
                    type_counts[DataType.INTEGER] += 1
                else:
                    type_counts[DataType.DECIMAL] += 1
            elif isinstance(value, datetime):
                type_counts[DataType.DATETIME] += 1
            elif isinstance(value, date):
                type_counts[DataType.DATE] += 1
            elif isinstance(value, time):
                type_counts[DataType.TIME] += 1
            else:
                # String or other types
                type_counts[DataType.TEXT] += 1
        
        # Determine primary type
        primary_type = max(type_counts, key=type_counts.get)
        
        # Generate original type description
        non_zero_types = {k: v for k, v in type_counts.items() if v > 0}
        original_type = f"excel_mixed({', '.join(f'{k.value}:{v}' for k, v in non_zero_types.items())})"
        
        return primary_type, original_type
    
    def _add_type_specific_stats(self, column: ColumnSchema, values: List, data_type: DataType):
        """Add statistics specific to the data type"""
        if not values:
            return
        
        if data_type in [DataType.INTEGER, DataType.DECIMAL, DataType.CURRENCY]:
            numeric_values = []
            for v in values:
                try:
                    if isinstance(v, (int, float)):
                        numeric_values.append(float(v))
                    else:
                        numeric_values.append(float(str(v).replace('$', '').replace(',', '')))
                except (ValueError, TypeError):
                    continue
            
            if numeric_values:
                column.min_value = min(numeric_values)
                column.max_value = max(numeric_values)
                column.avg_value = sum(numeric_values) / len(numeric_values)
        
        elif data_type == DataType.TEXT:
            str_values = [str(v) for v in values if v is not None]
            if str_values:
                lengths = [len(s) for s in str_values]
                column.min_length = min(lengths)
                column.max_length = max(lengths)
                column.avg_length = sum(lengths) / len(lengths)
    
    def _detect_column_constraints(self, column: ColumnSchema, values: List):
        """Detect column constraints and characteristics"""
        if not values:
            return
        
        # Check uniqueness
        str_values = [str(v) for v in values if v is not None]
        unique_values = set(str_values)
        
        if len(unique_values) == len(str_values) and column.null_count == 0:
            column.is_unique = True
            
        # Detect potential primary keys
        if column.is_unique and any(keyword in column.name.lower() for keyword in ['id', 'key', 'pk']):
            if not column.constraints:
                column.constraints = []
            column.constraints.append("PRIMARY_KEY_CANDIDATE")
        
        # Detect foreign keys
        if (column.data_type == DataType.IDENTIFIER and 
            not column.is_unique and 
            any(keyword in column.name.lower() for keyword in ['id', 'key', 'fk']) and
            'id' in column.name.lower() and column.name.lower() != 'id'):
            column.is_foreign_key = True
        
        # Detect categorical nature
        if len(unique_values) / len(str_values) < 0.1 and len(unique_values) < 20:
            column.data_type = DataType.CATEGORICAL
    
    def _determine_table_type(self, sheet_name: str, columns: List[ColumnSchema]) -> str:
        """Determine the type/purpose of the table"""
        sheet_lower = sheet_name.lower()
        
        # Check for common sheet patterns
        if any(pattern in sheet_lower for pattern in ['summary', 'dashboard', 'report']):
            return "summary_sheet"
        elif any(pattern in sheet_lower for pattern in ['data', 'raw', 'export']):
            return "data_sheet"
        elif any(pattern in sheet_lower for pattern in ['lookup', 'reference', 'master']):
            return "reference_sheet"
        elif len(columns) > 15:
            return "detailed_data_sheet"
        else:
            return "excel_sheet"
    
    def _generate_table_description(self, sheet_name: str, columns: List[ColumnSchema], row_count: int) -> str:
        """Generate description for the table"""
        id_cols = [col for col in columns if col.data_type == DataType.IDENTIFIER]
        currency_cols = [col for col in columns if col.data_type == DataType.CURRENCY]
        date_cols = [col for col in columns if col.data_type in [DataType.DATE, DataType.DATETIME]]
        
        desc_parts = [f"Excel sheet '{sheet_name}' with {len(columns)} columns and {row_count} data rows"]
        
        if id_cols:
            desc_parts.append(f"Contains identifiers: {', '.join(col.name for col in id_cols)}")
        
        if currency_cols and date_cols:
            desc_parts.append("Appears to contain transactional/financial data")
        elif currency_cols:
            desc_parts.append("Contains financial/monetary data")
        elif date_cols:
            desc_parts.append("Contains temporal data for trend analysis")
        
        return " | ".join(desc_parts)
    
    def _generate_column_description(self, name: str, data_type: DataType, values: List) -> str:
        """Generate description for individual columns"""
        name_lower = name.lower()
        
        # Business context descriptions
        if data_type == DataType.IDENTIFIER:
            if 'customer' in name_lower:
                return "Customer identifier for relationship tracking"
            elif 'order' in name_lower:
                return "Order identifier for transaction tracking" 
            elif 'product' in name_lower:
                return "Product identifier for inventory tracking"
            else:
                return "Unique identifier field"
        
        elif data_type == DataType.EMAIL:
            return "Email addresses for customer communication"
        
        elif data_type == DataType.CURRENCY:
            if 'price' in name_lower:
                return "Pricing information in monetary format"
            elif 'cost' in name_lower:
                return "Cost data in monetary format"
            elif 'total' in name_lower or 'amount' in name_lower:
                return "Total monetary amounts"
            else:
                return "Financial/monetary values"
        
        elif data_type == DataType.CATEGORICAL:
            unique_count = len(set(str(v) for v in values)) if values else 0
            if unique_count <= 5:
                return f"Categorical field with {unique_count} distinct values"
            else:
                return f"Classification field with {unique_count} categories"
        
        elif data_type in [DataType.DATE, DataType.DATETIME]:
            if 'created' in name_lower or 'date' in name_lower:
                return "Temporal data for chronological analysis"
            else:
                return "Date/time information for trend analysis"
        
        else:
            return f"{data_type.value.title()} data field"
    
    def _infer_business_context(self, tables: List[TableSchema]) -> str:
        """Infer the business context of the Excel file"""
        all_columns = []
        for table in tables:
            all_columns.extend(table.columns)
        
        # Count column types
        column_names = [col.name.lower() for col in all_columns]
        
        # Business domain detection
        if any('customer' in name for name in column_names):
            if any('order' in name or 'purchase' in name for name in column_names):
                return "customer_transaction_data"
            else:
                return "customer_data"
        
        elif any('product' in name for name in column_names):
            if any('inventory' in name or 'stock' in name for name in column_names):
                return "inventory_management"
            else:
                return "product_catalog"
        
        elif any('employee' in name or 'staff' in name for name in column_names):
            return "hr_employee_data"
        
        elif any('financial' in name or 'accounting' in name for name in column_names):
            return "financial_data"
        
        # Check for financial indicators
        currency_cols = [col for col in all_columns if col.data_type == DataType.CURRENCY]
        if len(currency_cols) >= 2:
            return "financial_analysis_data"
        
        # Default based on complexity
        if len(all_columns) > 20:
            return "complex_analytical_data"
        else:
            return "general_business_data"
    
    def get_source_type(self) -> str:
        return "xlsx"


# Helper function for content-based extraction (for refresh operations)
async def extract_xlsx_schema_from_content(file_content: bytes, filename: str = "excel_file.xlsx") -> DataSourceSchema:
    """Extract schema from Excel file content bytes"""
    
    class MockUploadFile:
        def __init__(self, content: bytes, filename: str):
            self.content = content
            self.filename = filename
            self.content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        async def read(self) -> bytes:
            return self.content
        
        async def seek(self, position: int = 0) -> None:
            pass
    
    mock_file = MockUploadFile(file_content, filename)
    extractor = XLSXSchemaExtractor()
    return await extractor.extract_schema(mock_file)


# Example usage and testing
if __name__ == "__main__":
    import asyncio
    
    async def test_xlsx_extractor():
        """Test the XLSX extractor"""
        print("XLSX Schema Extractor Test")
        print("=" * 50)
        
        # Note: In real usage, you'd have actual file upload
        print("To test this extractor:")
        print("1. Use it in your FastAPI endpoint with actual UploadFile")
        print("2. The extractor will return a unified DataSourceSchema")
        print("3. Call schema.to_llm_prompt() for LLM-optimized description")
        
        # Mock example of what the output structure would look like
        example_output = {
            "source_name": "sales_data.xlsx",
            "source_type": "xlsx",
            "total_tables": 2,
            "total_columns": 15,
            "tables": [
                {
                    "name": "Orders",
                    "table_type": "data_sheet",
                    "row_count": 1000,
                    "columns": [
                        {
                            "name": "order_id",
                            "data_type": "identifier",
                            "is_unique": True,
                            "sample_values": ["ORD001", "ORD002"],
                            "description": "Order identifier for transaction tracking"
                        },
                        {
                            "name": "total_amount", 
                            "data_type": "currency",
                            "min_value": 10.50,
                            "max_value": 999.99,
                            "avg_value": 156.75,
                            "description": "Total monetary amounts"
                        }
                    ]
                }
            ]
        }
        
        print("\nExample output structure:")
        print(json.dumps(example_output, indent=2))
    
    asyncio.run(test_xlsx_extractor())

